import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from openpyxl import load_workbook
import io

# Constants
INCH_CM = 2.54
RAPORT_16_9 = (16 / ((16**2 + 9**2)**0.5), 9 / ((16**2 + 9**2)**0.5))

def calculeaza_diagonala(d_m):
    return round(d_m * 39.37 * 0.84)

def dimensiuni_televizor(diagonala_inch):
    diagonala_cm = diagonala_inch * INCH_CM
    latime = diagonala_cm * RAPORT_16_9[0]
    inaltime = diagonala_cm * RAPORT_16_9[1]
    return round(latime / 100, 2), round(inaltime / 100, 2)

st.set_page_config(page_title="Configurare TV Kuziini", layout="wide")
components.html("<script>window.scrollTo(0, 0);</script>", height=0)

# ✅ Logo + titlu
st.markdown("""
<style>
    .logo-container {
        margin-top: -10px;
        margin-bottom: 0.7rem;
        text-align: center;
    }
    .logo-container img {
        width: 100%;
        max-width: 320px;
        height: auto;
    }
    .main-title {
        text-align: center;
        font-size: 1.7rem;
        margin-top: 0;
        margin-bottom: 1.5rem;
        color: black;
    }
    @media screen and (max-width: 768px) {
        .main-title {
            font-size: 1.4rem;
        }
    }
</style>
<div class='logo-container'>
    <st.image("Kuziini_logo_negru.png", use_column_width=False, width=320)>
</div>
<div class='main-title'>📐 Configurarea diagonalei TV în funcție de distanță</div>
""", unsafe_allow_html=True)

# Excel
excel_file_path = "www.xlsx"
wb = load_workbook(excel_file_path, data_only=True)
ws = wb.active

col1, col2 = st.columns([1, 1])

with col1:
    distanta = st.slider("📏 Alege distanța de vizionare (m)", min_value=1.0, max_value=10.0, value=2.5, step=0.1)
    diagonala_inch = calculeaza_diagonala(distanta)
    latime_m, inaltime_m = dimensiuni_televizor(diagonala_inch)

    st.markdown(f"""
    <style>
        .recomandare-box {{
            background-color:#F0F9FF;
            padding: 1.5rem 1rem;
            border-radius: 16px;
            border: 2px solid #0B5394;
            text-align: center;
            width: 100%;
            max-width: 480px;
            aspect-ratio: 9 / 5;
            margin: 0.5rem auto 1.2rem auto;
            display: flex;
            flex-direction: column;
            justify-content: center;
        }}
        .recomandare-box h1 {{
            font-size: 3rem;
            color: #FF5722;
            margin: 0;
        }}
        .recomandare-box h3 {{
            color: #0B5394;
            margin: 0.4rem;
        }}
        .recomandare-box p {{
            font-size: 1.3rem;
            margin: 0.2rem;
        }}
        @media (max-width: 768px) {{
            .recomandare-box {{
                padding: 1rem;
                aspect-ratio: 16 / 9;
            }}
        }}
    </style>
    <div class='recomandare-box'>
        <h1>{diagonala_inch}"</h1>
        <h3>Kuziini recomandă</h3>
        <p>pentru distanța de {distanta} metri</p>
        <p style='font-weight:bold;'>🖼️ {latime_m} m lățime × {inaltime_m} m înălțime</p>
    </div>
    """, unsafe_allow_html=True)

    if st.button("💾 Exportă în Excel cu aceste valori"):
        ws["B1"] = distanta
        ws["B6"] = round(distanta / 30, 2)
        ws["B7"] = round(distanta / 25, 2)
        output = io.BytesIO()
        wb.save(output)
        st.download_button(
            label="📥 Descarcă fișierul Excel actualizat",
            data=output.getvalue(),
            file_name="recomandare_tv_actualizat.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

with col2:
    st.image("TV.png", caption="Kuziini participa activ la inovatie", use_container_width=True)
    st.markdown("<p style='text-align:center;font-weight:bold;color:black;'>Living Kuziini × Samsung</p>", unsafe_allow_html=True)
